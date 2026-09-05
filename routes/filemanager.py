"""
Blueprint: File Manager routes
  GET  /api/fm/browse?path=<dir>   - list directory contents
  GET  /api/fm/read?path=<file>    - read file (json/yaml/env/csv/xlsx)
  POST /api/fm/save                - save file
"""
from __future__ import annotations

import csv
import io
import json
import sys
from pathlib import Path

from flask import Blueprint, jsonify, request
from ui_launcher.config_reader import ConfigReader
from ui_launcher.command_builder import resolve_python


def _ensure_repo_venv_on_path() -> None:
    """
    Add the active repo venv's site-packages to sys.path so packages like
    openpyxl that are installed only in the repo venv (not the tool's own
    Python) can be imported by the Flask server.
    Called lazily before any import that may live in the repo venv.
    """
    try:
        cfg      = ConfigReader().load()
        repo     = cfg.get("repo_root", "").strip()
        venv_cfg = cfg.get("venv_path", "").strip()
        if not repo:
            return
        python = resolve_python(repo, venv_cfg)
        if python == sys.executable:
            return   # tool and repo share the same Python — nothing to add
        # site-packages lives at <venv>/lib/pythonX.Y/site-packages (POSIX)
        # or <venv>/Lib/site-packages (Windows)
        venv_root = Path(python).parent.parent
        for sp in venv_root.rglob("site-packages"):
            sp_str = str(sp)
            if sp_str not in sys.path:
                sys.path.insert(0, sp_str)
            break
    except Exception:
        pass

bp = Blueprint("filemanager", __name__)

ALLOWED_EXTENSIONS = {".json", ".xlsx", ".xls", ".yaml", ".yml", ".env", ".csv"}

FILE_ICONS = {
    ".json":  "📋",
    ".yaml":  "📝",
    ".yml":   "📝",
    ".env":   "🔑",
    ".csv":   "📊",
    ".xlsx":  "📊",
    ".xls":   "📊",
}


def _safe_path(raw: str) -> Path | None:
    try:
        return Path(raw).resolve()
    except Exception:
        return None


def _ext(p: Path) -> str:
    name = p.name.lower()
    if name == ".env" or name.startswith(".env."):
        return ".env"
    return p.suffix.lower()


# ── Browse ─────────────────────────────────────────────────────────────────────

@bp.route("/api/fm/browse")
def browse():
    raw = request.args.get("path", "").strip() or str(Path.home())
    p = _safe_path(raw)
    if p is None or not p.exists():
        return jsonify({"error": "Path not found"}), 404
    if p.is_file():
        return jsonify({"error": "Not a directory"}), 400

    entries = []
    try:
        for child in sorted(p.iterdir(), key=lambda x: (x.is_file(), x.name.lower())):
            if child.name.startswith(".") and _ext(child) != ".env":
                continue
            ext = _ext(child)
            if child.is_dir():
                entries.append({"name": child.name, "path": str(child), "type": "dir"})
            elif ext in ALLOWED_EXTENSIONS:
                entries.append({
                    "name": child.name,
                    "path": str(child),
                    "type": "file",
                    "ext":  ext,
                    "icon": FILE_ICONS.get(ext, "📄"),
                })
    except PermissionError:
        return jsonify({"error": "Permission denied"}), 403

    return jsonify({
        "path":    str(p),
        "parent":  str(p.parent) if p != p.parent else None,
        "entries": entries,
    })


# ── Read ───────────────────────────────────────────────────────────────────────

@bp.route("/api/fm/read")
def read_file():
    raw = request.args.get("path", "").strip()
    p   = _safe_path(raw)
    if p is None or not p.is_file():
        return jsonify({"error": "File not found"}), 404

    ext = _ext(p)
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error": "Unsupported file type"}), 400

    # ── JSON ──────────────────────────────────────────────────────────────────
    if ext == ".json":
        try:
            content = p.read_text(encoding="utf-8")
            parsed  = json.loads(content)
            return jsonify({
                "type":    "json",
                "content": json.dumps(parsed, indent=2),
                "parsed":  parsed,
            })
        except Exception as exc:
            return jsonify({"error": f"Invalid JSON: {exc}"}), 400

    # ── YAML ──────────────────────────────────────────────────────────────────
    if ext in (".yaml", ".yml"):
        content = p.read_text(encoding="utf-8")
        try:
            _ensure_repo_venv_on_path()
            import yaml
            parsed = yaml.safe_load(content)
            return jsonify({
                "type":    "yaml",
                "content": content,
                "parsed":  parsed,
                "valid":   True,
            })
        except Exception as exc:
            return jsonify({
                "type":    "yaml",
                "content": content,
                "parsed":  None,
                "valid":   False,
                "error":   str(exc),
            })

    # ── ENV ───────────────────────────────────────────────────────────────────
    if ext == ".env":
        content = p.read_text(encoding="utf-8")
        pairs   = []
        for line in content.splitlines():
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                pairs.append({"comment": True, "raw": line})
                continue
            if "=" in stripped:
                k, _, v = stripped.partition("=")
                # strip surrounding quotes from value
                v = v.strip()
                if len(v) >= 2 and v[0] == v[-1] and v[0] in ('"', "'"):
                    v = v[1:-1]
                pairs.append({"comment": False, "key": k.strip(), "value": v})
            else:
                pairs.append({"comment": True, "raw": line})
        return jsonify({"type": "env", "content": content, "pairs": pairs})

    # ── CSV ───────────────────────────────────────────────────────────────────
    if ext == ".csv":
        try:
            content = p.read_text(encoding="utf-8-sig")  # handle BOM
            reader  = csv.reader(io.StringIO(content))
            rows    = [list(row) for row in reader]
            return jsonify({"type": "csv", "rows": rows})
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    # ── Excel ─────────────────────────────────────────────────────────────────
    if ext == ".xls":
        return jsonify({
            "error": "Old .xls format is not supported. Re-save as .xlsx in Excel first."
        }), 400

    try:
        _ensure_repo_venv_on_path()
        import openpyxl
        wb = openpyxl.load_workbook(str(p), data_only=True)
    except ImportError:
        return jsonify({"error": "openpyxl is not installed. Run: pip install openpyxl"}), 500
    except Exception as exc:
        msg = str(exc)
        if "not a zip" in msg.lower() or "badzip" in msg.lower():
            msg = "Cannot read — may be old .xls format or corrupted. Re-save as .xlsx."
        return jsonify({"error": msg}), 500

    try:
        sheets = {}
        for sn in wb.sheetnames:
            ws   = wb[sn]
            rows = [[("" if v is None else str(v)) for v in row]
                    for row in ws.iter_rows(values_only=True)]
            sheets[sn] = rows
        return jsonify({"type": "excel", "sheets": sheets, "sheet_names": wb.sheetnames})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ── Save ───────────────────────────────────────────────────────────────────────

@bp.route("/api/fm/save", methods=["POST"])
def save_file():
    body = request.json or {}
    raw  = body.get("path", "").strip()
    p    = _safe_path(raw)
    if p is None:
        return jsonify({"error": "Invalid path"}), 400

    ext = _ext(p)
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error": "Unsupported file type"}), 400

    # ── JSON ──────────────────────────────────────────────────────────────────
    if ext == ".json":
        content = body.get("content", "")
        try:
            parsed = json.loads(content)
        except json.JSONDecodeError as exc:
            return jsonify({"error": f"Invalid JSON: {exc}"}), 400
        try:
            p.write_text(json.dumps(parsed, indent=2), encoding="utf-8")
        except OSError as exc:
            return jsonify({"error": str(exc)}), 500
        return jsonify({"ok": True})

    # ── YAML ──────────────────────────────────────────────────────────────────
    if ext in (".yaml", ".yml"):
        content = body.get("content", "")
        try:
            _ensure_repo_venv_on_path()
            import yaml
            yaml.safe_load(content)
        except Exception as exc:
            return jsonify({"error": f"Invalid YAML: {exc}"}), 400
        try:
            p.write_text(content, encoding="utf-8")
        except OSError as exc:
            return jsonify({"error": str(exc)}), 500
        return jsonify({"ok": True})

    # ── ENV ───────────────────────────────────────────────────────────────────
    if ext == ".env":
        content = body.get("content", "")
        try:
            p.write_text(content, encoding="utf-8")
        except OSError as exc:
            return jsonify({"error": str(exc)}), 500
        return jsonify({"ok": True})

    # ── CSV ───────────────────────────────────────────────────────────────────
    if ext == ".csv":
        rows = body.get("rows", [])
        buf  = io.StringIO()
        csv.writer(buf).writerows(rows)
        try:
            p.write_text(buf.getvalue(), encoding="utf-8")
        except OSError as exc:
            return jsonify({"error": str(exc)}), 500
        return jsonify({"ok": True})

    # ── Excel ─────────────────────────────────────────────────────────────────
    try:
        _ensure_repo_venv_on_path()
        import openpyxl
        sheets_data = body.get("sheets", {})
        wb = openpyxl.load_workbook(str(p)) if p.exists() else openpyxl.Workbook()
        # remove default empty sheet if new workbook
        if "Sheet" in wb.sheetnames and not p.exists():
            del wb["Sheet"]
        for sn, rows in sheets_data.items():
            if sn in wb.sheetnames:
                ws = wb[sn]
                ws.delete_rows(1, ws.max_row)
            else:
                ws = wb.create_sheet(title=sn)
            for row in rows:
                ws.append(row)
        wb.save(str(p))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
    return jsonify({"ok": True})
